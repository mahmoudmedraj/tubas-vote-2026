import os, json, hashlib, threading, base64
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory
import time as _time

app = Flask(__name__, static_folder='static', static_url_path='')

# ── DB helpers ────────────────────────────────────────────────
_DB_LOCK = threading.Lock()

def get_db():
    url = os.environ.get('DATABASE_URL','')
    if not url: return None
    try:
        import psycopg2
        conn = psycopg2.connect(url, connect_timeout=5)
        return conn
    except: return None

def db_get(key, default=None):
    conn = get_db()
    if not conn:
        return CACHE.get(key, default)
    try:
        with conn.cursor() as c:
            c.execute("SELECT value FROM kv_store WHERE key=%s", (key,))
            row = c.fetchone()
            conn.close()
            return json.loads(row[0]) if row else default
    except:
        conn.close()
        return CACHE.get(key, default)

def db_set(key, value):
    CACHE[key] = value
    conn = get_db()
    if not conn: return
    try:
        with conn.cursor() as c:
            c.execute("""
                INSERT INTO kv_store(key,value) VALUES(%s,%s)
                ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value
            """, (key, json.dumps(value, ensure_ascii=False)))
        conn.commit(); conn.close()
    except Exception as e:
        print(f"[DB] set error: {e}"); conn.close()

def init_db():
    conn = get_db()
    if not conn: return
    try:
        with conn.cursor() as c:
            c.execute("""CREATE TABLE IF NOT EXISTS kv_store(
                key TEXT PRIMARY KEY, value TEXT)""")
        conn.commit(); conn.close()
        print("[DB] connected ✅")
    except Exception as e:
        print(f"[DB] init error: {e}")

CACHE = {}   # in-memory fallback

# ── In-memory duplicate protection ───────────────────────────
VOTED_CACHE   = set()
DEVICES_CACHE = {}
IP_CACHE      = {}

def h(s): return hashlib.sha256(str(s).encode()).hexdigest()[:16]

def load_data():
    voted = db_get('voted', {})
    for k in voted: VOTED_CACHE.add(k)
    print(f"[Data] loaded {len(VOTED_CACHE)} voted")

# ── Load voters & candidates ──────────────────────────────────
VOTERS_DB    = {}
CANDIDATES   = {}

def load_excel():
    paths = [
        'data/voters.xlsx',
        os.path.join(os.path.dirname(__file__), 'data', 'voters.xlsx')
    ]
    for path in paths:
        if not os.path.exists(path): continue
        try:
            import openpyxl
            wb  = openpyxl.load_workbook(path, read_only=True, data_only=True)

            # voters
            ws = wb['\u0633\u062c\u0644 \u0627\u0644\u0646\u0627\u062e\u0628\u064a\u0646']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[2]:
                    VOTERS_DB[str(row[2]).strip()] = str(row[3]).strip() if row[3] else ''

            # candidates
            ws2 = wb['\u0642\u0648\u0627\u0626\u0645 \u0627\u0644\u0645\u0631\u0634\u062d\u064a\u0646']
            for row in ws2.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                lid = str(row[0]).strip()
                if lid not in CANDIDATES:
                    CANDIDATES[lid] = {'name': str(row[1]).strip() if row[1] else '', 'members': []}
                CANDIDATES[lid]['members'].append({
                    'order': row[2], 'name': str(row[3]).strip() if row[3] else '',
                    'gender': str(row[4]).strip() if row[4] else '',
                    'status': str(row[5]).strip() if row[5] else ''
                })
            wb.close()
            print(f"[Excel] {len(VOTERS_DB)} voters, {len(CANDIDATES)} lists")
            return True
        except Exception as e:
            print(f"[Excel] error: {e}")
    return False

def find_voter(reg):
    return VOTERS_DB.get(str(reg).strip())

def record_device_vote(fp_h, ip_h, reg_h):
    DEVICES_CACHE[fp_h] = reg_h
    IP_CACHE[ip_h]      = IP_CACHE.get(ip_h, 0) + 1
    devs = db_get('devices', {})
    devs[fp_h] = reg_h
    db_set('devices', devs)

# ── Routes ────────────────────────────────────────────────────
@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/api/voter', methods=['POST'])
def api_voter():
    data = request.get_json(silent=True) or {}
    reg  = str(data.get('reg_num','')).strip()
    if not reg:
        return jsonify({'ok':False,'error':'أدخل رقم التسجيل الانتخابي'}), 400
    name = find_voter(reg)
    if name is None:
        return jsonify({'ok':False,'error':'رقم التسجيل غير موجود في السجل الانتخابي'}), 404
    reg_h = h(reg)
    # Check already voted
    voted = db_get('voted', {})
    if reg_h in voted or reg_h in VOTED_CACHE:
        return jsonify({'ok':False,'error':'لقد قمت بالتصويت مسبقاً، لا يمكن التصويت مرتين'}), 403
    # Check device
    fp  = str(data.get('fp','')).strip()
    fp_h = h(fp) if fp else None
    if fp_h:
        devs = db_get('devices', {})
        if fp_h in devs or fp_h in DEVICES_CACHE:
            return jsonify({'ok':False,'error':'هذا الجهاز استخدم للتصويت مسبقاً'}), 403
    return jsonify({'ok':True,'name':name})

@app.route('/api/candidates', methods=['GET'])
def api_candidates():
    out = []
    for lid, info in CANDIDATES.items():
        out.append({'id':lid,'name':info['name'],'members':info['members']})
    return jsonify({'ok':True,'lists':out})

@app.route('/api/vote', methods=['POST'])
def api_vote():
    data     = request.get_json(silent=True) or {}
    reg      = str(data.get('reg_num','')).strip()
    list_id  = str(data.get('list_id','')).strip()
    chosen   = data.get('candidates', [])
    fp       = str(data.get('fp','')).strip()
    ip       = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()

    if not reg or not list_id:
        return jsonify({'ok':False,'error':'بيانات ناقصة'}), 400

    reg_h = h(reg); fp_h = h(fp); ip_h = h(ip)

    with _DB_LOCK:
        # Double-check: already voted?
        voted = db_get('voted', {})
        if reg_h in voted or reg_h in VOTED_CACHE:
            return jsonify({'ok':False,'error':'لقد صوّتت مسبقاً'}), 403
        # Double-check: device?
        devs = db_get('devices', {})
        if fp_h in devs or fp_h in DEVICES_CACHE:
            return jsonify({'ok':False,'error':'هذا الجهاز استخدم للتصويت مسبقاً'}), 403

        # Record vote
        votes = db_get('votes', {'total':0,'lists':{},'candidates':{}})
        votes['total'] = votes.get('total',0) + 1
        votes['lists'][list_id] = votes['lists'].get(list_id,0) + 1
        for c in chosen:
            votes['candidates'][c] = votes['candidates'].get(c,0) + 1
        db_set('votes', votes)

        # Mark voter
        voted[reg_h] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        VOTED_CACHE.add(reg_h)
        db_set('voted', voted)

        # Mark device
        record_device_vote(fp_h, ip_h, reg_h)

    return jsonify({'ok':True,'msg':'تم تسجيل رأيك بنجاح'})

# ── Admin ─────────────────────────────────────────────────────
ADMIN_PW = lambda: os.environ.get('ADMIN_PW','Tubas@0598652625')

@app.route('/api/admin/login', methods=['POST'])
def api_admin_login():
    data = request.get_json(silent=True) or {}
    if data.get('password') == ADMIN_PW():
        return jsonify({'ok':True})
    return jsonify({'ok':False,'error':'كلمة المرور غير صحيحة'}), 401

@app.route('/api/admin/results', methods=['POST'])
def api_admin_results():
    data = request.get_json(silent=True) or {}
    if data.get('password') != ADMIN_PW():
        return jsonify({'ok':False,'error':'غير مصرح'}), 401
    votes   = db_get('votes',   {'total':0,'lists':{},'candidates':{}})
    voted   = db_get('voted',   {})
    devices = db_get('devices', {})
    return jsonify({'ok':True,'votes':votes,
                    'total_voters':len(voted),
                    'unique_devices':len(devices)})

@app.route('/api/admin/toggle', methods=['POST'])
def api_admin_toggle():
    data = request.get_json(silent=True) or {}
    if data.get('password') != ADMIN_PW():
        return jsonify({'ok':False}), 401
    cur = db_get('open', True)
    db_set('open', not cur)
    return jsonify({'ok':True,'open': not cur})

@app.route('/api/status')
def api_status():
    return jsonify({'open': db_get('open', True)})

@app.route('/api/debug')
def api_debug():
    return jsonify({
        'voters': len(VOTERS_DB),
        'lists':  len(CANDIDATES),
        'voted':  len(db_get('voted',{})),
        'devices':len(db_get('devices',{})),
        'db_connected': get_db() is not None,
    })

# ── Start ─────────────────────────────────────────────────────
load_excel()
load_data()
init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
